# coding=utf-8

from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl import worksheet
import os

def readFromXLSX(path):
    wb = load_workbook(path)
    ws = wb.active

    ws_rows = tuple(ws.rows)
    rows_len = len(ws_rows)

    return rows_len-1, ws_rows[1:rows_len]

def writeAssignment(form, asn_name):
    save_path = 'Assignment_' + asn_name +'_Form.xlsx'
    if os.path.exists(save_path):
        os.remove(save_path)

    wb = Workbook()
    ws = wb.active

    row_num = len(form)

    ws['A1'] = '团队ID'
    ws['B1'] = '团队名称'
    ws['C1'] = '作业提交情况'
    ws['D1'] = '作业分数'

    for i in range(row_num):
        ws['A'+str(i+2)] = form[i][0]
        ws['B'+str(i+2)] = form[i][1]
        ws['C'+str(i+2)] = form[i][2]
        ws['D'+str(i+2)] = form[i][3]

    wb.save(save_path)

    return save_path

def writeAllAssignment(form, course_name):
    save_path ='Course_' + course_name + '_Assignment_Form.xlsx'
    if os.path.exists(save_path):
        os.remove(save_path)

    wb = Workbook()
    ws = wb.active
    now = 0

    for team in form:
        row_num = len(team)
        if now > 0:
            ws = wb.create_sheet()
        ws.title = team[0][1]

        for i in range(row_num):
            ws['A'+str(i+1)] = team[i][0]
            ws['B'+str(i+1)] = team[i][1]
            ws['C'+str(i+1)] = team[i][2]
            ws['D'+str(i+1)] = team[i][3]

        now = now + 1

    wb.save(save_path)

    return save_path

def writeTeam(form, course_name):
    save_path = 'Course_' + course_name +'_Team_Form.xlsx'
    if os.path.exists(save_path):
        os.remove(save_path)

    wb = Workbook()
    ws = wb.active
    now = 0

    for team in form:
        if now > 0:
            ws = wb.create_sheet()
        ws.title = team[0][1]
        row_num = len(team)

        ws['A1'] = '团队ID'
        ws['B1'] = '团队名称'
        ws['C1'] = '学生学号'
        ws['D1'] = '学生姓名'
        ws['E1'] = '担任角色'

        for i in range(row_num):
            ws['A'+str(i+2)] = team[i][0]
            ws['B'+str(i+2)] = team[i][1]
            ws['C'+str(i+2)] = team[i][2]
            ws['D'+str(i+2)] = team[i][3]
            ws['E'+str(i+2)] = team[i][4]

    wb.save(save_path)
    return save_path

def writeGrade(form, course_name):
    save_path = 'Course_' + course_name + '_Student_Grade_Form.xlsx'
    if os.path.exists(save_path):
        os.remove(save_path)

    wb = Workbook()
    ws = wb.active
    num = len(form)

    ws['A1'] = '学生学号'
    ws['B1'] = '学生姓名'
    ws['C1'] = '总成绩'

    for i in range(num):
        ws['A'+str(i+1)] = form[i][0]
        ws['B'+str(i+1)] = form[i][1]
        ws['C'+str(i+1)] = form[i][2]

    wb.save(save_path)
    return save_path

def fileSystemResponse(Resources, Folders):
    R1 = '<li class="myfile"><input type="text" class="changename" name="1" value="'
    R2 = '"/><input class="checkbox" name="'
    R3 = '" type="checkbox" value="" /></li>'
    R_str = []
    for res in Resources:
        R_str.append(R1 + res.name + R2 + str(res.id) + R3 + '\n')

    F1 = '<li class="myfolder"><input type="text" class="changename" name="1" value="'
    F2 = '"/><input class="checkbox" name="'
    F3 = '" type="checkbox" value="" /></li>'
    F_str = []
    for f in Folders:
        F_str.append(F1 + f.name + F2 + str(f.id) + F3 + '\n')

    ret_str = ''
    for res in R_str:
        ret_str = ret_str + res
    for f in F_str:
        ret_str = ret_str + f

    return ret_str

def deleteResource(path):
    for res_path in path:
        print(res_path)


if __name__ == '__main__':
    form1 = [[1, 'storm', '已提交', 100],
             [2, 'aaa', '未提交', 0]]
    xlsx1 = writeAssignment(form1, '测试')
    print(xlsx1)

    form2 = [[1, 'storm', '', ''],
             ['作业ID', '作业名称', '作业提交情况', '作业分数'],
             [1, '分析', '已提交', 98],
             [2, '开发', '已提交', 99],
             [3, '测试', '已提交', 100],
             [2, 'aaa', '', ''],
             ['作业ID', '作业名称', '作业提交情况', '作业分数'],
             [1, '分析', '未提交', 0],
             [2, '开发', '未提交', 0],
             [3, '测试', '未提交', 0]]
    xlsx2 = writeAllAssignment(form2, '软工过程')
    print(xlsx2)